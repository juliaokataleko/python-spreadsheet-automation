import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

product_per_supplier = {}
total_value_per_supllier = {}
products_under_10_inventory = {}

# get products number per supplier
for product_row in range(2, product_list.max_row + 1):
    # supplier name
    supplier_position = product_list.cell(product_row, 4)
    supplier_name = supplier_position.value
    
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row,3).value
    product_number = product_list.cell(product_row,1).value
    # add cell to file
    inventory_price = product_list.cell(product_row,5)
    
    # calculate number of products for supplier
    # check if supplier exists
    if supplier_name in product_per_supplier:
        # current_number_of_products = product_per_supplier[supplier_name]
        current_number_of_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_number_of_products + 1
    else:
        # print(f"Adding a new supplier in list {supplier_name}")
        product_per_supplier[supplier_name] = 1
        
    # calculate total inventory value per supplier
    if supplier_name in total_value_per_supllier:
        current_total_value = total_value_per_supllier[supplier_name]
        total_value_per_supllier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supllier[supplier_name] = inventory * price
        
    # products with inventory under 10
    if inventory < 10:
        products_under_10_inventory[product_number] = int(inventory)
        
    # login to add value for total inventory price
    inventory_price.value = inventory * price
    
inventory_file.save("inventory_with_total_value.xlsx")
              
print(product_per_supplier)
print(total_value_per_supllier)
print(products_under_10_inventory)
